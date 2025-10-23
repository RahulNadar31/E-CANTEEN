from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory
import os
import json
from werkzeug.utils import secure_filename
from database import init_db, get_db_connection

import csv
from datetime import datetime
import hmac
import hashlib
try:
    import razorpay  # optional; app runs without it
except ImportError:
    razorpay = None
try:
    # For PDF invoice generation
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib import colors
except Exception:
    # We'll handle missing dependency at runtime when generating PDF
    A4 = None
    canvas = None
    mm = None
    colors = None

app = Flask(__name__)
app.secret_key = 'smart_canteen_secret_key_2024'
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024
app.config['UPI_COLLECT_ID'] = '9769519263@ptaxis'
app.config['RAZORPAY_KEY_ID'] = os.environ.get('RAZORPAY_KEY_ID', '')
app.config['RAZORPAY_KEY_SECRET'] = os.environ.get('RAZORPAY_KEY_SECRET', '')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Custom Jinja2 filter for JSON parsing
@app.template_filter('from_json')
def from_json_filter(json_string):
    try:
        return json.loads(json_string)
    except (json.JSONDecodeError, TypeError):
        return []

# Database initialization - FIXED for newer Flask versions
with app.app_context():
    init_db()
  
        conn = get_db_connection()
        conn.execute('''
            INSERT OR REPLACE INTO kitchen_staff (staff_id, name, email, password, role, created_at)
            VALUES (1, 'Manasvi Kharpuse', 'manasvikharpuse2006@gmail.com', '12345678', 'kitchen', COALESCE((SELECT created_at FROM kitchen_staff WHERE staff_id=1), CURRENT_TIMESTAMP))
        ''')
        conn.commit()
        conn.close()
    except Exception:
        pass

@app.route('/')
def index():
    return render_template('index.html')

# Serve uploaded files (ID photos)
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    # Normalize Windows backslashes in stored paths
    safe_filename = filename.replace('\\', '/')
    return send_from_directory(app.config['UPLOADS_ROOT'], safe_filename)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user_type = request.form['user_type']
        
        if user_type == 'admin':
            user = verify_admin_login(email, password)
            if user:
                session['user_id'] = user['admin_id']
                session['user_type'] = 'admin'
                session['user_name'] = user['name']
                return redirect(url_for('admin_dashboard'))
            els'] = staff['staff_id']
                session['user_type'] = 'kitchen'
                session['user_name'] = staff['name']
                return redirect(url_for('kitchen_dashboard'))
            else:
                flash('Invalid kitchen credentials', 'danger')
        else:
            user = verify_user_login(email, password)
            if user:
                session['user_id'] = user['id']
                session['user_type'] = 'user'
                session['user_name'] = user['name']
                return redirect(url_for('student_dashboard'))
            else:
                flash('Invalid credentials or account not verified', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        pnr = request.form['pnr']
        password = request.form['password']
        
        id_photo_path = None
        if 'id_photo' in request.files:
            file = request.files['id_photo']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(f"{pnr}_{name}.{file.filename.rsplit('.', 1)[1].lower()}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                # Store relative path under uploads/ for portability
                id_photo_path = os.path.join('id_proofs', filename).replace('\\', '/')
        
        if register_user(name, email, pnr, password, id_photo_path):
            flash('Registration successful! Wait for admin verification.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Email already exists!', 'danger')
    
    return render_template('register.html')

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get stats
    total_users = conn.execute('SELECT COUNT(*) FROM users WHERE verified = 1').fetchone()[0]
    pending_verifications = conn.execute('SELECT COUNT(*) FROM users WHERE verified = 0').fetchone()[0]
    total_orders = conn.execute('SELECT COUNT(*) FROM orders').fetchone()[0]
    total_revenue = conn.execute('SELECT COALESCE(SUM(total_amount), 0) FROM orders WHERE payment_status = "Paid"').fetchone()[0]
    total_expenses = conn.execute('SELECT COALESCE(SUM(amount), 0) FROM expenses').fetchone()[0]
    profit = total_revenue - total_expenses
    
    # Get recent orders
    recent_orders = conn.execute('''
        SELECT o.*, u.name FROM orders o 
        JOIN users u ON o.user_id = u.id 
        ORDER BY o.order_time DESC LIMIT 5
    ''').fetchall()
    
    conn.close()
    
    return render_template('admin_dashboard.html', 
                         total_users=total_users,
                         pending_verifications=pending_verifications,
                         total_orders=total_orders,
                         total_revenue=total_revenue,
                         total_expenses=total_expenses,
                         profit=profit,
                         recent_orders=recent_orders)

@app.route('/admin/export')
def admin_export():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))

    export_type = request.args.get('type', 'xlsx')

    conn = get_db_connection()
    users = conn.execute('SELECT id, name, email, pnr, verified, id_photo_path, created_at FROM users').fetchall()
    orders = conn.execute('''
        SELECT o.*, u.name AS customer_name, u.email AS customer_email
        FROM orders o
        JOIN users u ON u.id = o.user_id
        ORDER BY o.order_time DESC
    ''').fetchall()
    expenses = conn.execute('SELECT * FROM expenses ORDER BY expense_date DESC').fetchall()
    conn.close()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Build absolute photo URL and stringify date/time to avoid Excel ######
    base_url = request.host_url.rstrip('/')

    users_header = ['id', 'name', 'email', 'pnr', 'verified', 'id_photo_url', 'created_at']
    users_rows = []
    for u in users:
        photo_rel = (u['id_photo_path'] or '').replace('\\', '/')
        # Strip any accidental leading 'uploads/' so we don't get uploads/uploads/...
        if photo_rel.startswith('uploads/'):
            photo_rel = photo_rel[len('uploads/'):]
        if photo_rel.startswith('/'):
            photo_rel = photo_rel[1:]
        photo_url = f"{base_url}/uploads/{photo_rel}" if photo_rel else ''
        users_rows.append([
            u['id'], u['name'], u['email'], u['pnr'], u['verified'], photo_url, str(u['created_at'])
        ])

    orders_header = [
        'order_id', 'customer_name', 'customer_email',
        'items_count', 'items_detail', 'items_json',
        'total_amount', 'payment_status', 'order_status',
        'order_time', 'estimated_time', 'preparation_started', 'preparation_completed', 'notification_sent'
    ]
    orders_rows = []
    for o in orders:
        items_json = o['items'] if isinstance(o['items'], str) else json.dumps(o['items'] or [])
        try:
            parsed_items = json.loads(items_json)
        except Exception:
            parsed_items = []
        items_count = 0
        details_parts = []
        for it in parsed_items:
            qty = int(it.get('quantity', 0) or 0)
            price = float(it.get('price', 0) or 0)
            name = str(it.get('name', 'Item'))
            items_count += qty
            details_parts.append(f"{name} x{qty} @ {price} = {price * qty}")
        items_detail = '; '.join(details_parts)
        orders_rows.append([
            o['order_id'], o['customer_name'], o['customer_email'],
            items_count, items_detail, items_json,
            o['total_amount'], o['payment_status'], o['order_status'],
            str(o['order_time']), o['estimated_time'], str(o['preparation_started']), str(o['preparation_completed']), o['notification_sent']
        ])

    expenses_header = ['expense_id', 'description', 'amount', 'category', 'expense_date', 'created_at']
    expenses_rows = []
    for e in expenses:
        expenses_rows.append([
            e['expense_id'], e['description'], e['amount'], e['category'], str(e['expense_date']), str(e['created_at'])
        ])

    if export_type == 'csv':
        # Multi-CSV in a simple text stream separated by headers
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['USERS'])
        writer.writerow(users_header)
        writer.writerows(users_rows)
        writer.writerow([])
        writer.writerow(['ORDERS'])
        writer.writerow(orders_header)
        writer.writerows(orders_rows)
        writer.writerow([])
        writer.writerow(['EXPENSES'])
        writer.writerow(expenses_header)
        writer.writerows(expenses_rows)
        output.seek(0)
        return app.response_class(output.getvalue(), mimetype='text/csv', headers={
            'Content-Disposition': f'attachment; filename="cafnova_export_{timestamp}.csv"'
        })
    else:
        # XLSX with three sheets
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            flash('Excel export requires openpyxl. Please install dependencies.', 'danger')
            return redirect(url_for('admin_dashboard'))

        wb = Workbook()
        # Users sheet
        ws_users = wb.active
        ws_users.title = 'Users'
        ws_users.append(users_header)
        for row in users_rows:
            ws_users.append(["" if v is None else v for v in row])
        # Orders sheet
        ws_orders = wb.create_sheet('Orders')
        ws_orders.append(orders_header)
        for row in orders_rows:
            ws_orders.append(["" if v is None else v for v in row])
        # Expenses sheet
        ws_expenses = wb.create_sheet('Expenses')
        ws_expenses.append(expenses_header)
        for row in expenses_rows:
            ws_expenses.append(["" if v is None else v for v in row])

        # Styling helpers
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='DDDDDD'),
                             right=Side(style='thin', color='DDDDDD'),
                             top=Side(style='thin', color='DDDDDD'),
                             bottom=Side(style='thin', color='DDDDDD'))

        def style_sheet(ws, header_row_index: int = 1):
            # Header styling
            for cell in ws[header_row_index]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            # Freeze header and enable filter
            ws.freeze_panes = ws['A2']
            ws.auto_filter.ref = ws.dimensions
            # Zebra striping and borders
            for idx, row in enumerate(ws.iter_rows(min_row=header_row_index+1, max_row=ws.max_row), start=0):
                if idx % 2 == 0:
                    for cell in row:
                        cell.fill = alt_fill
                for cell in row:
                    cell.border = thin_border

        style_sheet(ws_users)
        style_sheet(ws_orders)
        style_sheet(ws_expenses)

        # Number formats
        def apply_number_formats(ws, header_labels_to_formats: dict):
            header_indices = {cell.value: cell.column for cell in ws[1] if cell.value in header_labels_to_formats}
            for label, col_idx in header_indices.items():
                number_format = header_labels_to_formats[label]
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row)[0]:
                    cell.number_format = number_format

        # Apply formats: currency and integers
        inr_currency = '[$₹-409] #,##0.00'
        integer_format = '#,##0'
        apply_number_formats(ws_users, {})
        apply_number_formats(ws_orders, {
            'items_count': integer_format,
            'total_amount': inr_currency
        })
        apply_number_formats(ws_expenses, {
            'amount': inr_currency
        })

        # Basic formatting to avoid ###### in Excel
        # Set number_format to text for datetime-like columns and widen columns
        def set_text_and_width(ws, headers, text_cols, width_map=None):
            if width_map is None:
                width_map = {}
            # Header row is 1
            for idx, header in enumerate(headers, start=1):
                col_letter = get_column_letter(idx)
                # Set width
                ws.column_dimensions[col_letter].width = width_map.get(header, 20)
                # Apply text format for specific columns
                if header in text_cols:
                    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                        for cell in row:
                            cell.number_format = '@'

        set_text_and_width(
            ws_users,
            users_header,
            text_cols={'created_at', 'id_photo_url'},
            width_map={'id_photo_url': 60, 'created_at': 22, 'email': 28}
        )

        set_text_and_width(
            ws_orders,
            orders_header,
            text_cols={'items_json', 'items_detail', 'order_time', 'preparation_started', 'preparation_completed'},
            width_map={'items_json': 60, 'items_detail': 60, 'order_time': 22, 'preparation_started': 22, 'preparation_completed': 22, 'customer_email': 28}
        )

        set_text_and_width(
            ws_expenses,
            expenses_header,
            text_cols={'expense_date', 'created_at'},
            width_map={'description': 28, 'expense_date': 22, 'created_at': 22}
        )

        xlsx_stream = io.BytesIO()
        wb.save(xlsx_stream)
        xlsx_stream.seek(0)
        return app.response_class(xlsx_stream.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
            'Content-Disposition': f'attachment; filename="cafnova_export_{timestamp}.xlsx"'
        })

@app.route('/admin/verify-users')
def verify_users():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    pending_users = get_pending_verifications()
    return render_template('user_verification.html', pending_users=pending_users)

@app.route('/admin/verify-user/<int:user_id>')
def approve_user(user_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    verify_user(user_id)
    flash('User verified successfully!', 'success')
    return redirect(url_for('verify_users'))

@app.route('/admin/reject-user/<int:user_id>')
def reject_user_route(user_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    reject_user(user_id)
    flash('User rejected and removed!', 'success')
    return redirect(url_for('verify_users'))

@app.route('/admin/menu')
def menu_management():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    menu_items = conn.execute('SELECT * FROM menu ORDER BY category, item_name').fetchall()
    conn.close()
    
    return render_template('menu_management.html', menu_items=menu_items)

@app.route('/admin/add-menu-item', methods=['POST'])
def add_menu_item():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    item_name = request.form['item_name']
    price = float(request.form['price'])
    category = request.form['category']
    description = request.form['description']
    
    conn = get_db_connection()
    conn.execute('''
        INSERT INTO menu (item_name, price, category, description)
        VALUES (?, ?, ?, ?)
    ''', (item_name, price, category, description))
    conn.commit()
    conn.close()
    
    flash('Menu item added successfully!', 'success')
    return redirect(url_for('menu_management'))

@app.route('/admin/edit-menu-item/<int:item_id>', methods=['POST'])
def edit_menu_item(item_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    item_name = request.form['item_name']
    price = float(request.form['price'])
    category = request.form['category']
    description = request.form['description']
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE menu SET item_name = ?, price = ?, category = ?, description = ?
        WHERE item_id = ?
    ''', (item_name, price, category, description, item_id))
    conn.commit()
    conn.close()
    
    flash('Menu item updated successfully!', 'success')
    return redirect(url_for('menu_management'))

@app.route('/admin/delete-menu-item/<int:item_id>')
def delete_menu_item(item_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    conn.execute('DELETE FROM menu WHERE item_id = ?', (item_id,))
    conn.commit()
    conn.close()
    
    flash('Menu item deleted successfully!', 'success')
    return redirect(url_for('menu_management'))

@app.route('/admin/expenses')
def expense_management():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    expenses = conn.execute('SELECT * FROM expenses ORDER BY expense_date DESC').fetchall()
    
    # Calculate total expenses
    total_expenses = conn.execute('SELECT COALESCE(SUM(amount), 0) FROM expenses').fetchone()[0]
    
    # Calculate total revenue
    total_revenue = conn.execute('SELECT COALESCE(SUM(total_amount), 0) FROM orders WHERE payment_status = "Paid"').fetchone()[0]
    
    # Calculate profit
    profit = total_revenue - total_expenses
    
    conn.close()
    
    return render_template('expense_management.html', 
                         expenses=expenses,
                         total_expenses=total_expenses,
                         total_revenue=total_revenue,
                         profit=profit)

@app.route('/admin/add-expense', methods=['POST'])
def add_expense():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    description = request.form['description']
    amount = float(request.form['amount'])
    category = request.form['category']
    
    conn = get_db_connection()
    conn.execute('''
        INSERT INTO expenses (description, amount, category)
        VALUES (?, ?, ?)
    ''', (description, amount, category))
    conn.commit()
    conn.close()
    
    flash('Expense added successfully!', 'success')
    return redirect(url_for('expense_management'))

@app.route('/admin/delete-expense/<int:expense_id>')
def delete_expense(expense_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    conn.execute('DELETE FROM expenses WHERE expense_id = ?', (expense_id,))
    conn.commit()
    conn.close()
    
    flash('Expense deleted successfully!', 'success')
    return redirect(url_for('expense_management'))

@app.route('/student/dashboard')
def student_dashboard():
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    menu_items = conn.execute('SELECT * FROM menu WHERE available = 1').fetchall()
    
    # Get user orders
    user_orders = conn.execute('''
        SELECT * FROM orders WHERE user_id = ? ORDER BY order_time DESC
    ''', (session['user_id'],)).fetchall()
    
    conn.close()
    
    # Build serializable orders with parsed items
    user_orders_processed = []
    for row in user_orders:
        order_dict = dict(row)
        try:
            raw_items = order_dict.get('items', '[]')
            order_dict['items_parsed'] = json.loads(raw_items) if isinstance(raw_items, str) else (raw_items or [])
        except Exception:
            order_dict['items_parsed'] = []
        user_orders_processed.append(order_dict)

    return render_template('student_dashboard.html', 
                         menu_items=menu_items, 
                         user_orders=user_orders_processed)

@app.route('/place-order', methods=['POST'])
def place_order():
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))
    
    # Validate and compute total on server
    raw_items = request.form.get('items', '[]')
    try:
        cart_items = json.loads(raw_items)
        if not isinstance(cart_items, list):
            cart_items = []
    except Exception:
        cart_items = []

    if len(cart_items) == 0:
        flash('Your cart is empty or invalid.', 'danger')
        return redirect(url_for('student_dashboard'))

    # Fetch current prices from DB and compute total
    conn = get_db_connection()
    item_map = {}
    for row in conn.execute('SELECT item_id, item_name, price FROM menu'):
        item_map[row['item_id']] = {'name': row['item_name'], 'price': float(row['price'])}

    normalized_items = []
    total_amount = 0.0
    for it in cart_items:
        try:
            item_id = int(it.get('id'))
            quantity = int(it.get('quantity', 1))
            if item_id in item_map and quantity > 0:
                name = item_map[item_id]['name']
                price = item_map[item_id]['price']
                normalized_items.append({
                    'id': item_id,
                    'name': name,
                    'price': price,
                    'quantity': quantity
                })
                total_amount += price * quantity
        except Exception:
            continue

    if len(normalized_items) == 0:
        conn.close()
        flash('Invalid items in cart. Please try again.', 'danger')
        return redirect(url_for('student_dashboard'))

    items_json = json.dumps(normalized_items)

    # Store order with Pending payment
    cursor = conn.execute('''
        INSERT INTO orders (user_id, items, total_amount, payment_status)
        VALUES (?, ?, ?, 'Pending')
    ''', (session['user_id'], items_json, total_amount))
    order_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Redirect to payment page
    return redirect(url_for('payment_page', order_id=order_id))

@app.route('/payment/<int:order_id>')
def payment_page(order_id):
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    order = conn.execute('SELECT * FROM orders WHERE order_id = ? AND user_id = ?', (order_id, session['user_id'])).fetchone()
    conn.close()
    
    if not order:
        flash('Order not found!', 'danger')
        return redirect(url_for('student_dashboard'))
    
    # Build dict with parsed items (avoid mutating sqlite Row)
    order_dict = dict(order)
    try:
        raw_items = order_dict.get('items', '[]')
        order_dict['items_parsed'] = json.loads(raw_items) if isinstance(raw_items, str) else (raw_items or [])
    except Exception:
        order_dict['items_parsed'] = []
    
    # Build UPI deeplink for display (simulated)
    upi_id = app.config.get('UPI_COLLECT_ID', '')
    amount = order_dict.get('total_amount', 0)
    upi_pay_url = f"upi://pay?pa={upi_id}&pn=Cafnova&am={amount}&cu=INR" if upi_id else ''
    
    # Razorpay order create (only if keys are configured)
    rzp_key = app.config['RAZORPAY_KEY_ID']
    rzp_secret = app.config['RAZORPAY_KEY_SECRET']
    rzp_order = None
    if rzp_key and rzp_secret and razorpay is not None:
        try:
            client = razorpay.Client(auth=(rzp_key, rzp_secret))
            rzp_order = client.order.create(dict(
                amount=int(float(order_dict.get('total_amount', 0)) * 100),
                currency='INR',
                receipt=f"cafnova_{order_id}",
                payment_capture=1
            ))
        except Exception:
            rzp_order = None
    
    return render_template('payment.html', order=order_dict, upi_collect_id=upi_id, upi_pay_url=upi_pay_url, rzp_key=rzp_key, rzp_order=rzp_order)

@app.route('/payment/verify', methods=['POST'])
def payment_verify():
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))

    rzp_key = app.config['RAZORPAY_KEY_ID']
    rzp_secret = app.config['RAZORPAY_KEY_SECRET']
    if not (rzp_key and rzp_secret) or razorpay is None:
        flash('Payment verification not configured', 'danger')
        return redirect(url_for('student_dashboard'))

    payload = request.form
    order_id = payload.get('order_id')
    rzp_payment_id = payload.get('razorpay_payment_id')
    rzp_order_id = payload.get('razorpay_order_id')
    rzp_signature = payload.get('razorpay_signature')

    # Verify signature
    generated_signature = hmac.new(
        rzp_secret.encode('utf-8'),
        f"{rzp_order_id}|{rzp_payment_id}".encode('utf-8'),
        hashlib.sha256
    ).hexdigest()

    if hmac.compare_digest(generated_signature, rzp_signature or ''):
        conn = get_db_connection()
        conn.execute('UPDATE orders SET payment_status = "Paid" WHERE order_id = ?', (order_id,))
        conn.commit()
        conn.close()
        flash('Payment successful! Your order has been confirmed.', 'success')
        return redirect(url_for('download_invoice_pdf', order_id=order_id))
    else:
        flash('Payment verification failed', 'danger')
        return redirect(url_for('payment_page', order_id=order_id))

@app.route('/webhook/razorpay', methods=['POST'])
def webhook_razorpay():
    rzp_secret = app.config['RAZORPAY_KEY_SECRET']
    if not rzp_secret:
        return '', 200
    # Optional: verify webhook signature header 'X-Razorpay-Signature'
    event = request.get_json(silent=True) or {}
    if event.get('event') == 'payment.captured':
        payload = event.get('payload', {})
        notes = payload.get('payment', {}).get('entity', {}).get('notes', {})
        order_id = notes.get('cafnova_order_id')
        if order_id:
            conn = get_db_connection()
            conn.execute('UPDATE orders SET payment_status = "Paid" WHERE order_id = ?', (order_id,))
            conn.commit()
            conn.close()
    return '', 200

@app.route('/process-payment/<int:order_id>', methods=['POST'])
def process_payment(order_id):
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))
    
    payment_method = request.form.get('payment_method')
    card_number = request.form.get('card_number', '')
    expiry_date = request.form.get('expiry_date', '')
    cvv = request.form.get('cvv', '')
    upi_id = request.form.get('upi_id', '')
    
    # Simulate payment processing (like Dominoz)
    if payment_method == 'card' and card_number and expiry_date and cvv:
        # Simulate successful card payment
        conn = get_db_connection()
        conn.execute('''
            UPDATE orders SET payment_status = 'Paid' WHERE order_id = ?
        ''', (order_id,))
        conn.commit()
        conn.close()
        flash('Payment successful! Your order has been confirmed.', 'success')
        return redirect(url_for('download_invoice_pdf', order_id=order_id))
    elif payment_method == 'upi' and upi_id:
        # Simulate successful UPI payment
        conn = get_db_connection()
        conn.execute('''
            UPDATE orders SET payment_status = 'Paid' WHERE order_id = ?
        ''', (order_id,))
        conn.commit()
        conn.close()
        flash('UPI payment successful! Your order has been confirmed.', 'success')
        return redirect(url_for('download_invoice_pdf', order_id=order_id))
    else:
        flash('Payment failed! Please check your details.', 'danger')
        return redirect(url_for('payment_page', order_id=order_id))

@app.route('/admin/update-order-status/<int:order_id>', methods=['POST'])
def update_order_status(order_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return {'success': False, 'message': 'Unauthorized'}, 401
    
    data = request.get_json()
    new_status = data.get('status')
    
    if new_status not in ['Pending', 'Preparing', 'Ready']:
        return {'success': False, 'message': 'Invalid status'}, 400
    
    conn = get_db_connection()
    # Allow updating status only for Paid orders
    order = conn.execute('SELECT payment_status FROM orders WHERE order_id = ?', (order_id,)).fetchone()
    if not order or order['payment_status'] != 'Paid':
        conn.close()
        return {'success': False, 'message': 'Order is not paid. Cannot update status.'}, 400

    conn.execute(
        'UPDATE orders SET order_status = ? WHERE order_id = ?',
        (new_status, order_id)
    )
    conn.commit()
    conn.close()
    
    return {'success': True, 'message': 'Order status updated'}

# Assume payment: marks an order as Paid and takes user to invoice
@app.route('/assume-paid/<int:order_id>', methods=['POST'])
def assume_paid(order_id):
    if 'user_type' not in session or session['user_type'] != 'user':
        return redirect(url_for('login'))

    conn = get_db_connection()
    order = conn.execute('SELECT user_id, payment_status FROM orders WHERE order_id = ?', (order_id,)).fetchone()
    if not order or order['user_id'] != session['user_id']:
        conn.close()
        flash('Order not found.', 'danger')
        return redirect(url_for('student_dashboard'))

    if order['payment_status'] != 'Paid':
        conn.execute("UPDATE orders SET payment_status = 'Paid' WHERE order_id = ?", (order_id,))
        conn.commit()
    conn.close()

    flash('Payment marked as successful. Generating invoice...', 'success')
    return redirect(url_for('download_invoice_pdf', order_id=order_id))

# Generate and download invoice PDF for an order
@app.route('/invoice/<int:order_id>.pdf')
def download_invoice_pdf(order_id):
    if 'user_type' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    order = conn.execute('SELECT * FROM orders WHERE order_id = ?', (order_id,)).fetchone()
    user_row = None
    if order:
        user_row = conn.execute('SELECT * FROM users WHERE id = ?', (order['user_id'],)).fetchone()
    conn.close()

    # Permission: user can download own invoice; admin/kitchen can download any
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('student_dashboard' if session.get('user_type') == 'user' else 'kitchen_dashboard'))
    if session.get('user_type') == 'user' and order['user_id'] != session.get('user_id'):
        flash('Unauthorized.', 'danger')
        return redirect(url_for('student_dashboard'))

    # Parse items
    try:
        items = json.loads(order['items']) if isinstance(order['items'], str) else (order['items'] or [])
    except Exception:
        items = []

    # If reportlab not available
    if canvas is None or A4 is None:
        flash('PDF generation dependency missing. Please install ReportLab.', 'danger')
        return redirect(url_for('student_dashboard'))

    # Build PDF in-memory
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Header
    pdf.setFillColor(colors.black)
    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(20 * mm, height - 20 * mm, 'Cafnova - Tax Invoice')
    pdf.setFont('Helvetica', 10)
    pdf.drawString(20 * mm, height - 26 * mm, f"Order ID: #{order['order_id']}")
    pdf.drawString(20 * mm, height - 31 * mm, f"Order Time: {str(order['order_time'])[:19]}")
    if user_row:
        pdf.drawString(20 * mm, height - 36 * mm, f"Customer: {user_row['name']} ({user_row['email']})")

    # Table header
    start_y = height - 50 * mm
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(20 * mm, start_y, 'Item')
    pdf.drawRightString(140 * mm, start_y, 'Qty')
    pdf.drawRightString(170 * mm, start_y, 'Amount (₹)')
    pdf.line(20 * mm, start_y - 3 * mm, 190 * mm, start_y - 3 * mm)

    # Table rows
    y = start_y - 10 * mm
    pdf.setFont('Helvetica', 10)
    total = 0.0
    for it in items:
        name = str(it.get('name', 'Item'))
        qty = int(it.get('quantity', 0) or 0)
        price = float(it.get('price', 0) or 0)
        amount = qty * price
        total += amount
        pdf.drawString(20 * mm, y, name)
        pdf.drawRightString(140 * mm, y, str(qty))
        pdf.drawRightString(170 * mm, y, f"{amount:.2f}")
        y -= 7 * mm
        if y < 30 * mm:
            pdf.showPage()
            y = height - 20 * mm

    # Total
    pdf.setFont('Helvetica-Bold', 12)
    pdf.line(120 * mm, y - 2 * mm, 190 * mm, y - 2 * mm)
    pdf.drawRightString(170 * mm, y - 8 * mm, f"Total: ₹{float(order['total_amount']):.2f}")

    pdf.setFont('Helvetica-Oblique', 9)
    pdf.drawString(20 * mm, 15 * mm, 'Thank you for ordering with Cafnova!')

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return app.response_class(buffer.getvalue(), mimetype='application/pdf', headers={
        'Content-Disposition': f"attachment; filename=invoice_{order_id}.pdf"
    })

@app.route('/kitchen/dashboard')
def kitchen_dashboard():
    if 'user_type' not in session or session['user_type'] != 'kitchen':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get pending orders (only Paid orders are visible to kitchen)
    pending_orders = conn.execute('''
        SELECT o.*, u.name as customer_name FROM orders o 
        JOIN users u ON o.user_id = u.id 
        WHERE o.order_status = 'Pending' AND o.payment_status = 'Paid'
        ORDER BY o.order_time ASC
    ''').fetchall()
    
    # Get preparing orders
    preparing_orders = conn.execute('''
        SELECT o.*, u.name as customer_name FROM orders o 
        JOIN users u ON o.user_id = u.id 
        WHERE o.order_status = 'Preparing' AND o.payment_status = 'Paid'
        ORDER BY o.order_time ASC
    ''').fetchall()
    
    # Get ready orders
    ready_orders = conn.execute('''
        SELECT o.*, u.name as customer_name FROM orders o 
        JOIN users u ON o.user_id = u.id 
        WHERE o.order_status = 'Ready' AND o.payment_status = 'Paid'
        ORDER BY o.order_id DESC
    ''').fetchall()
    
    # Compute top-selling items (across all Paid orders)
    top_map = {}
    all_paid = conn.execute('SELECT items FROM orders WHERE payment_status = "Paid"').fetchall()
    for r in all_paid:
        try:
            parsed = json.loads(r['items']) if isinstance(r['items'], str) else (r['items'] or [])
        except Exception:
            parsed = []
        for it in parsed:
            name = str(it.get('name', 'Item'))
            qty = int(it.get('quantity', 0) or 0)
            top_map[name] = top_map.get(name, 0) + qty
    # Build top list sorted
    top_sellers = sorted(({ 'name': n, 'quantity': q } for n, q in top_map.items()), key=lambda x: x['quantity'], reverse=True)[:5]

    conn.close()
    
    # Build lists with parsed items
    def process_rows(rows):
        out = []
        for r in rows:
            d = dict(r)
            try:
                raw_items = d.get('items', '[]')
                d['items_parsed'] = json.loads(raw_items) if isinstance(raw_items, str) else (raw_items or [])
            except Exception:
                d['items_parsed'] = []
            out.append(d)
        return out

    return render_template('kitchen_dashboard.html', 
                         pending_orders=process_rows(pending_orders),
                         preparing_orders=process_rows(preparing_orders),
                         ready_orders=process_rows(ready_orders),
                         top_sellers=top_sellers)

@app.route('/kitchen/start-preparation/<int:order_id>')
def start_preparation(order_id):
    if 'user_type' not in session or session['user_type'] != 'kitchen':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    # Ensure order is Paid before starting preparation
    order = conn.execute('SELECT payment_status FROM orders WHERE order_id = ?', (order_id,)).fetchone()
    if not order or order['payment_status'] != 'Paid':
        conn.close()
        flash('Cannot start preparation. Payment not confirmed.', 'danger')
        return redirect(url_for('kitchen_dashboard'))

    conn.execute('''
        UPDATE orders SET order_status = 'Preparing', preparation_started = CURRENT_TIMESTAMP
        WHERE order_id = ?
    ''', (order_id,))
    conn.commit()
    conn.close()
    
    flash('Order preparation started!', 'success')
    return redirect(url_for('kitchen_dashboard'))

@app.route('/kitchen/complete-order/<int:order_id>')
def complete_order(order_id):
    if 'user_type' not in session or session['user_type'] != 'kitchen':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE orders SET order_status = 'Ready', preparation_completed = CURRENT_TIMESTAMP, notification_sent = 1
        WHERE order_id = ?
    ''', (order_id,))
    conn.commit()
    conn.close()
    
    flash('Order completed and notification sent!', 'success')
    return redirect(url_for('kitchen_dashboard'))

@app.route('/kitchen/set-time/<int:order_id>', methods=['POST'])
def set_preparation_time(order_id):
    if 'user_type' not in session or session['user_type'] != 'kitchen':
        return redirect(url_for('login'))
    
    estimated_time = request.form.get('estimated_time', 15)
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE orders SET estimated_time = ?
        WHERE order_id = ?
    ''', (estimated_time, order_id))
    conn.commit()
    conn.close()
    
    flash(f'Preparation time set to {estimated_time} minutes', 'success')
    return redirect(url_for('kitchen_dashboard'))

@app.route('/kitchen/check-new-orders')
def check_new_orders():
    if 'user_type' not in session or session['user_type'] != 'kitchen':
        return {'success': False, 'message': 'Unauthorized'}, 401
    
    conn = get_db_connection()
    
    # Count new pending orders (Paid but not started)
    new_orders_count = conn.execute('''
        SELECT COUNT(*) as count FROM orders 
        WHERE order_status = 'Pending' AND payment_status = 'Paid'
    ''').fetchone()['count']
    
    # Get preparing orders that might be overdue
    preparing_orders = conn.execute('''
        SELECT order_id, estimated_time, preparation_started,
               (julianday('now') - julianday(preparation_started)) * 24 * 60 as minutes_elapsed
        FROM orders 
        WHERE order_status = 'Preparing' AND payment_status = 'Paid'
    ''').fetchall()
    
    overdue_orders = []
    for order in preparing_orders:
        if order['minutes_elapsed'] and order['minutes_elapsed'] > order['estimated_time']:
            overdue_orders.append(order['order_id'])
    
    conn.close()
    
    return {
        'success': True,
        'new_orders': new_orders_count,
        'overdue_orders': overdue_orders,
        'timestamp': datetime.now().isoformat()
    }

@app.route('/api/order-status/<int:order_id>')
def get_order_status(order_id):
    if 'user_type' not in session or session['user_type'] != 'user':
        return {'success': False, 'message': 'Unauthorized'}, 401
    
    conn = get_db_connection()
    order = conn.execute('''
        SELECT order_id, order_status, payment_status, estimated_time, 
               preparation_started, preparation_completed,
               (julianday('now') - julianday(preparation_started)) * 24 * 60 as minutes_elapsed
        FROM orders 
        WHERE order_id = ? AND user_id = ?
    ''', (order_id, session['user_id'])).fetchone()
    
    if not order:
        conn.close()
        return {'success': False, 'message': 'Order not found'}, 404
    
    # Calculate progress if order is being prepared
    progress = 0
    remaining_time = order['estimated_time']
    
    if order['order_status'] == 'Preparing' and order['preparation_started']:
        elapsed = order['minutes_elapsed'] or 0
        progress = min((elapsed / order['estimated_time']) * 100, 100) if order['estimated_time'] > 0 else 0
        remaining_time = max(order['estimated_time'] - elapsed, 0)
    
    conn.close()
    
    return {
        'success': True,
        'order_id': order['order_id'],
        'status': order['order_status'],
        'payment_status': order['payment_status'],
        'estimated_time': order['estimated_time'],
        'progress': round(progress, 1),
        'remaining_time': round(remaining_time, 1),
        'preparation_started': order['preparation_started'],
        'preparation_completed': order['preparation_completed']
    }

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully!', 'info')
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)